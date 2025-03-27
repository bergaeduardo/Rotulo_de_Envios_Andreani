import openpyxl
import win32com.client as win32
import os
import logging

logger = logging.getLogger(__name__)

def copy_range(start_col, start_row, end_col, end_row, sheet):
    """
    Copia un rango de celdas de una hoja de cálculo de Excel.

    Args:
        start_col: Columna inicial del rango.
        start_row: Fila inicial del rango.
        end_col: Columna final del rango.
        end_row: Fila final del rango.
        sheet: Hoja de cálculo de origen.

    Returns:
        Una lista anidada con los valores del rango copiado.
    """
    range_selected = []
    for i in range(start_row, end_row + 1):
        row_selected = []
        for j in range(start_col, end_col + 1):
            row_selected.append(sheet.cell(row=i, column=j).value)
        range_selected.append(row_selected)
    return range_selected

def paste_range(start_col, start_row, end_col, end_row, sheet_receiving, copied_data):
    """
    Pega datos en un rango de celdas de una hoja de cálculo de Excel.

    Args:
        start_col: Columna inicial del rango de destino.
        start_row: Fila inicial del rango de destino.
        end_col: Columna final del rango de destino.
        end_row: Fila final del rango de destino.
        sheet_receiving: Hoja de cálculo de destino.
        copied_data: Lista anidada con los datos a pegar.
    """
    count_row = 0
    for i in range(start_row, end_row + 1):
        count_col = 0
        for j in range(start_col, end_col + 1):
            try:
                sheet_receiving.cell(row=i, column=j).value = copied_data[count_row][count_col]
            except IndexError as e:
                logger.error(f"IndexError al pegar en celda ({i}, {j}): {e}")
                break # Romper el bucle interno si copied_data es más corto que el rango de destino
            count_col += 1
        count_row += 1

def clear_cells(template_excel_path, cio=1, start_row=3, end_col=29):
    """
    Limpia el contenido de un rango de celdas en una plantilla de Excel.

    Args:
        template_excel_path: Ruta al archivo Excel plantilla.
        start_col: Columna inicial del rango a limpiar.
        start_row: Fila inicial del rango a limpiar.
        end_col: Columna final del rango a limpiar.
    """
    try:
        template = openpyxl.load_workbook(template_excel_path, read_only=False)
        temp_sheet = template.active
        ii=3
        while not(temp_sheet.cell(row=ii, column=1).value == None):
            while cio <= 29:
                if cio != 2:
                    temp_sheet.cell(row=ii, column=cio).value = ''
                    cio+=1
                else:
                    cio+=1
            cio=1
            ii+=1
        template.save(template_excel_path)
        logger.info(f"Plantilla Excel '{template_excel_path}' limpiada.")
    except FileNotFoundError:
        logger.error(f"Error: No se encontró la plantilla Excel en la ruta: {template_excel_path}")
        raise
    except Exception as e:
        logger.error(f"Error al limpiar la plantilla Excel '{template_excel_path}': {e}")
        raise

def save_excel_xlsx_format(excel_file_path):
    """
    Guarda un archivo Excel en formato .xlsx usando win32com.

    Args:
        excel_file_path: Ruta al archivo Excel.
    """
    try:
        excel_app = win32.gencache.EnsureDispatch('Excel.Application')
        excel_app.Visible = False  # No mostrar Excel
        excel_app.DisplayAlerts = False # No mostrar alertas

        workbook = excel_app.Workbooks.Open(excel_file_path)
        workbook.SaveAs(excel_file_path, FileFormat=51)  # 51 para .xlsx
        workbook.Close()
        excel_app.DisplayAlerts = True
        excel_app.Quit()
        logger.info(f"Archivo Excel guardado en formato .xlsx: '{excel_file_path}'")
    except Exception as e:
        logger.error(f"Error al guardar el archivo Excel en formato .xlsx '{excel_file_path}': {e}")
        raise

def load_excel_workbook(file_path):
    """
    Carga un libro de trabajo de Excel usando openpyxl.

    Args:
        file_path: Ruta al archivo Excel.

    Returns:
        Un objeto Workbook de openpyxl.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        logger.info(f"Libro de trabajo Excel cargado: '{file_path}'")
        return workbook
    except FileNotFoundError:
        logger.error(f"Error: Archivo Excel no encontrado en la ruta: {file_path}")
        raise
    except Exception as e:
        logger.error(f"Error al cargar el libro de trabajo Excel '{file_path}': {e}")
        raise

if __name__ == '__main__':
    # Ejemplo de uso (necesitarás archivos Excel de prueba)
    logging.basicConfig(level=logging.INFO)

    # Crear archivos Excel de prueba si no existen
    test_input_file = 'test_input.xlsx'
    test_template_file = 'test_template.xlsx'

    if not os.path.exists(test_input_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Data 1'
        sheet['B1'] = 'Data 2'
        wb.save(test_input_file)

    if not os.path.exists(test_template_file):
        wb = openpyxl.Workbook()
        wb.save(test_template_file)


    workbook_origin = load_excel_workbook(test_input_file)
    sheet_origin = workbook_origin.active
    workbook_dest = load_excel_workbook(test_template_file)
    sheet_dest = workbook_dest.active

    data_to_copy = copy_range(1, 1, 2, 1, sheet_origin) # Copiar A1:B1
    paste_range(1, 1, 2, 1, sheet_dest, data_to_copy) # Pegar en A1:B1 de template
    workbook_dest.save('test_template_updated.xlsx')
    print("Rango copiado y pegado. Ver 'test_template_updated.xlsx'")

    clear_cells(test_template_file)
    print(f"Plantilla '{test_template_file}' limpiada.")

    save_excel_xlsx_format('test_template_updated.xlsx')
    print("Archivo 'test_template_updated.xlsx' guardado en formato .xlsx")
